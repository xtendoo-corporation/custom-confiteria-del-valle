import logging
from io import BytesIO

import openpyxl

_logger = logging.getLogger(__name__)


class TodopinturaExcelParser:
    """Parseador de archivos Excel de Todopintura (UMACLI)."""

    # Mapeo de columnas: nombre en Excel -> índice (1-based)
    COLUMN_MAP = {
        "numero_cliente": 1,
        "nombre_cliente": 2,
        "direccion": 3,
        "codigo_postal": 4,
        "telefono_1": 5,
        "telefono_2": 6,
        "dni_cif": 7,
        "recargo_equivalencia": 8,
        "forma_pago": 9,
        "necesita_vale": 10,
        "comercial": 11,
        "limite_credito": 12,
        "posibilidad_mo": 13,
        "estado": 14,
        "iban": 15,
        "tienda_cobro": 16,
        "rino": 17,
        "descuento_fijo": 18,
        "pronto_pago": 19,
        "observaciones_1": 20,
        "observaciones_2": 21,
        "observaciones_3": 22,
        "observaciones_4": 23,
        "correo_electronico": 24,
        "nombre_contacto": 25,
        "condiciones_pago": 26,
        "metodo_pago": 27,
    }

    def __init__(self, file_content):
        """
        Inicializa el parser con el contenido del archivo Excel.

        Args:
            file_content: bytes del archivo Excel
        """
        self.file_content = file_content
        self.workbook = None
        self.worksheet = None
        self.errors = []
        self.warnings = []
        self.total_records = 0
        self.processed_records = []

    def parse(self):
        """
        Parsea el archivo Excel y retorna una lista de diccionarios con los datos.

        Returns:
            list: Lista de diccionarios con los datos del cliente
        """
        try:
            self.workbook = openpyxl.load_workbook(
                BytesIO(self.file_content), data_only=True
            )
            self.worksheet = self.workbook.active

            if self.worksheet.max_row < 2:
                self.errors.append("El archivo Excel está vacío o no contiene datos.")
                return []

            records = []
            seen_keys = set()  # Para detectar duplicados

            # Iterar desde la fila 2 (saltar encabezado)
            for row_idx in range(2, self.worksheet.max_row + 1):
                try:
                    record = self._parse_row(row_idx)
                    if record:
                        # Detectar duplicados por NUMERO CLIENTE
                        external_id = record.get("external_id")
                        if external_id in seen_keys:
                            self.warnings.append(
                                f"Fila {row_idx}: Duplicado de cliente {external_id}. "
                                "Se utilizará el último registro encontrado."
                            )
                            # Remover el registro anterior
                            records = [
                                r
                                for r in records
                                if r.get("external_id") != external_id
                            ]

                        seen_keys.add(external_id)
                        records.append(record)
                        self.total_records += 1
                except Exception as e:
                    self.errors.append(f"Fila {row_idx}: {str(e)}")
                    _logger.warning(f"Error parsing row {row_idx}: {e}")

            self.processed_records = records
            return records

        except Exception as e:
            self.errors.append(f"Error al leer el archivo Excel: {str(e)}")
            _logger.error(f"Error parsing Excel file: {e}")
            return []

    def _parse_row(self, row_idx):
        """
        Parsea una fila individual del Excel.

        Args:
            row_idx: índice de la fila (1-based)

        Returns:
            dict: Diccionario con los datos del cliente
        """
        ws = self.worksheet

        def get_cell_value(col_num, default=None):
            """Obtiene el valor de una celda de forma segura."""
            try:
                cell = ws.cell(row=row_idx, column=col_num)
                value = cell.value
                return value if value is not None else default
            except Exception:
                return default

        # Campos básicos
        numero_cliente = get_cell_value(self.COLUMN_MAP["numero_cliente"])
        nombre_cliente = get_cell_value(self.COLUMN_MAP["nombre_cliente"], "").strip()

        # Validar campos obligatorios
        if not numero_cliente or not nombre_cliente:
            return None

        numero_cliente = str(numero_cliente).strip()

        # Parsear todos los campos
        record = {
            "external_id": numero_cliente,
            "name": nombre_cliente,
            "street": get_cell_value(self.COLUMN_MAP["direccion"], "").strip(),
            "zip": str(get_cell_value(self.COLUMN_MAP["codigo_postal"], "")).strip(),
            "phone": str(get_cell_value(self.COLUMN_MAP["telefono_1"], "")).strip(),
            "mobile": str(get_cell_value(self.COLUMN_MAP["telefono_2"], "")).strip(),
            "vat": str(get_cell_value(self.COLUMN_MAP["dni_cif"], "")).strip(),
            "email": str(
                get_cell_value(self.COLUMN_MAP["correo_electronico"], "")
            ).strip(),
            "contact_name": get_cell_value(
                self.COLUMN_MAP["nombre_contacto"], ""
            ).strip(),
            "payment_term": get_cell_value(
                self.COLUMN_MAP["condiciones_pago"], ""
            ).strip(),
            "payment_method": get_cell_value(
                self.COLUMN_MAP["metodo_pago"], ""
            ).strip(),
            "comercial_code": get_cell_value(self.COLUMN_MAP["comercial"]),
            "row_number": row_idx,
        }

        # Campos numéricos
        try:
            limite_credito = get_cell_value(self.COLUMN_MAP["limite_credito"])
            record["credit_limit"] = float(limite_credito) if limite_credito else 0.0
        except (ValueError, TypeError):
            record["credit_limit"] = 0.0

        try:
            descuento_fijo = get_cell_value(self.COLUMN_MAP["descuento_fijo"])
            record["discount_fixed"] = float(descuento_fijo) if descuento_fijo else 0.0
        except (ValueError, TypeError):
            record["discount_fixed"] = 0.0

        try:
            pronto_pago = get_cell_value(self.COLUMN_MAP["pronto_pago"])
            record["discount_early_payment"] = (
                float(pronto_pago) if pronto_pago else 0.0
            )
        except (ValueError, TypeError):
            record["discount_early_payment"] = 0.0

        # Booleano
        necesita_vale = (
            get_cell_value(self.COLUMN_MAP["necesita_vale"], "").strip().upper()
        )
        record["needs_voucher"] = "yes" if necesita_vale == "S" else "no"

        # Observaciones concatenadas
        obs_parts = [
            get_cell_value(self.COLUMN_MAP["observaciones_1"], "").strip(),
            get_cell_value(self.COLUMN_MAP["observaciones_2"], "").strip(),
            get_cell_value(self.COLUMN_MAP["observaciones_3"], "").strip(),
            get_cell_value(self.COLUMN_MAP["observaciones_4"], "").strip(),
        ]
        record["comment"] = "\n".join([o for o in obs_parts if o])

        # Limpiar valores vacíos
        for key in ["phone", "mobile", "email", "vat", "zip", "street"]:
            if record[key] == "0" or record[key] == "":
                record[key] = None

        return record

    def get_summary(self):
        """
        Retorna un resumen del parseo.

        Returns:
            dict: Resumen con estadísticas
        """
        return {
            "total_records": self.total_records,
            "errors": self.errors,
            "warnings": self.warnings,
            "success": len(self.errors) == 0,
        }
