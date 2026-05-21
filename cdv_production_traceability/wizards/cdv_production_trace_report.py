from odoo import api, fields, models, _
from odoo.exceptions import UserError
from datetime import datetime, date
import base64
import io


class CdvProductionTraceReport(models.TransientModel):
    _name = "cdv.production.trace.report"
    _description = "Asistente de informe de trazabilidad"

    date_from = fields.Date(
        string="Fecha desde",
        required=True,
        default=lambda self: date(date.today().year, 1, 1),
    )
    date_to = fields.Date(
        string="Fecha hasta",
        required=True,
        default=fields.Date.context_today,
    )
    product_id = fields.Many2one(
        comodel_name="product.product",
        string="Producto terminado",
        domain=[("is_storable", "=", True)],
        required=True,
        help="Producto elaborado del que se desea conocer la trazabilidad",
    )
    lot_id = fields.Many2one(
        comodel_name="stock.lot",
        string="Lote",
        required=True,
        help="Lote del producto elaborado",
    )
    company_id = fields.Many2one(
        comodel_name="res.company",
        string="Compañía",
        required=True,
        default=lambda self: self.env.company,
    )
    warning_message = fields.Html(
        string="Mensaje de advertencia",
        readonly=True,
    )
    line_ids = fields.One2many(
        comodel_name="cdv.production.trace.report.line",
        inverse_name="wizard_id",
        string="Líneas de trazabilidad",
        readonly=True,
    )
    total_productions = fields.Integer(
        string="Total de producciones",
        compute="_compute_totals",
        store=False,
    )
    total_components = fields.Integer(
        string="Total de componentes rastreados",
        compute="_compute_totals",
        store=False,
    )
    total_qty_produced = fields.Float(
        string="Total de unidades producidas",
        compute="_compute_totals",
        store=False,
    )
    total_qty_uom_name = fields.Char(
        string="Unidad de medida",
        compute="_compute_totals",
        store=False,
    )

    @api.depends("line_ids")
    def _compute_totals(self):
        """Calcular totales para el resumen"""
        for record in self:
            # Contar producciones únicas (producto + lote)
            productions = record.line_ids.mapped(lambda l: (l.finished_product_id.id, l.finished_lot_id.id))
            record.total_productions = len(set(productions))
            # Contar componentes únicos encontrados
            record.total_components = len(record.line_ids.filtered(lambda l: l.trace_status == 'found'))
            # Sumar total de unidades producidas
            # Agrupamos por producto y lote para no contar duplicados
            qty_dict = {}
            for line in record.line_ids:
                key = (line.finished_product_id.id, line.finished_lot_id.id)
                if key not in qty_dict:
                    qty_dict[key] = line.finished_qty
            record.total_qty_produced = sum(qty_dict.values())
            # Obtener el nombre de la unidad
            uom_names = set(record.line_ids.mapped('finished_uom_id.name'))
            record.total_qty_uom_name = ", ".join(filter(None, uom_names))

    @api.onchange("product_id")
    def _onchange_product_id(self):
        """Limpiar lote cuando cambia el producto"""
        if self.product_id:
            self.lot_id = False
            return {
                "domain": {
                    "lot_id": [
                        ("product_id", "=", self.product_id.id),
                        ("company_id", "=", self.company_id.id),
                    ]
                }
            }
        return {"domain": {"lot_id": []}}

    def action_compute_trace(self):
        """Calcular trazabilidad"""
        self.ensure_one()

        if self.date_from > self.date_to:
            raise UserError(
                _("La fecha desde debe ser menor o igual a la fecha hasta.")
            )

        # Limpiar líneas anteriores y mensaje de advertencia
        self.line_ids.unlink()
        self.warning_message = False

        # Buscar partes de producción en el rango
        domain = [
            ("is_production_entry", "=", True),
            ("state", "=", "done"),
            ("date_done", ">=", self.date_from),
            ("date_done", "<=", self.date_to),
            ("company_id", "=", self.company_id.id),
        ]

        production_entries = self.env["stock.picking"].search(domain)

        if not production_entries:
            raise UserError(
                _("No se encontraron partes de producción en el período seleccionado.")
            )

        # Procesar cada línea de producción
        lines_to_create = []

        for entry in production_entries:
            for line in entry.move_ids:
                # Filtrar por producto si está especificado
                if self.product_id and line.product_id != self.product_id:
                    continue

                # Obtener los lotes del movimiento
                move_lots = line.move_line_ids.mapped('lot_id')

                # Filtrar por lote si está especificado
                if self.lot_id and self.lot_id not in move_lots:
                    continue

                # Obtener BoM del producto
                boms = self.env["mrp.bom"]._bom_find(
                    line.product_id,
                    company_id=self.company_id.id,
                    bom_type="normal",
                )
                bom = boms[line.product_id]

                if not bom:
                    # Producto sin BoM - crear línea sin componentes
                    for move_line in line.move_line_ids:
                        lines_to_create.append(
                            {
                                "wizard_id": self.id,
                                "production_date": entry.date_done.date() if entry.date_done else entry.scheduled_date,
                                "finished_product_id": line.product_id.id,
                                "finished_lot_id": move_line.lot_id.id if move_line.lot_id else False,
                                "finished_qty": move_line.quantity,
                                "finished_uom_id": line.product_uom.id,
                                "component_product_id": False,
                                "component_qty_standard": 0.0,
                                "component_lot_id": False,
                                "trace_status": "missing",
                            }
                        )
                    continue

                # Procesar cada componente de la BoM
                bom_obj = bom

                for bom_line in bom_obj.bom_line_ids:
                    component_product = bom_line.product_id

                    # Convertir Date a Datetime para comparacion robusta
                    prod_date = entry.date_done.date() if entry.date_done else entry.scheduled_date
                    entry_date_start = datetime.combine(prod_date, datetime.min.time())
                    entry_date_end = datetime.combine(prod_date, datetime.max.time())

                    search_domain = [
                        ("product_id", "=", component_product.id),
                        ("company_id", "=", self.company_id.id),
                        ("date_from", "<=", entry_date_end),
                        "|",
                        ("date_to", "=", False),
                        ("date_to", ">=", entry_date_start),
                    ]

                    # Buscar TODAS las materias primas en uso en la fecha de producción
                    # Usamos active_test=False para encontrar también materias primas finalizadas (inactivas)
                    raw_materials = (
                        self.env["cdv.raw.material.in.use"]
                        .with_context(active_test=False)
                        .search(search_domain)
                    )

                    if not raw_materials:
                        # No se encontró ninguna materia prima en uso - crear línea sin componente
                        for move_line in line.move_line_ids:
                            lines_to_create.append(
                                {
                                    "wizard_id": self.id,
                                    "production_date": prod_date,
                                    "finished_product_id": line.product_id.id,
                                    "finished_lot_id": move_line.lot_id.id if move_line.lot_id else False,
                                    "finished_qty": move_line.quantity,
                                    "finished_uom_id": line.product_uom.id,
                                    "component_product_id": component_product.id,
                                    "component_qty_standard": bom_line.product_qty,
                                    "component_lot_id": False,
                                    "trace_status": "missing",
                                }
                            )
                    else:
                        # Crear una línea por cada lote del producto terminado y por cada lote de materia prima
                        for move_line in line.move_line_ids:
                            for raw_material in raw_materials:
                                lines_to_create.append(
                                    {
                                        "wizard_id": self.id,
                                        "production_date": prod_date,
                                        "finished_product_id": line.product_id.id,
                                        "finished_lot_id": move_line.lot_id.id if move_line.lot_id else False,
                                        "finished_qty": move_line.quantity,
                                        "finished_uom_id": line.product_uom.id,
                                        "component_product_id": component_product.id,
                                        "component_qty_standard": bom_line.product_qty,
                                        "component_lot_id": raw_material.lot_id.id if raw_material.lot_id else False,
                                        "trace_status": "found",
                                    }
                                )

        if lines_to_create:
            self.env["cdv.production.trace.report.line"].create(lines_to_create)
            self.warning_message = False
        else:
            # Mostrar mensaje de advertencia en el wizard en lugar de popup
            self.warning_message = """
                <h4 class="alert-heading">
                    <i class="fa fa-exclamation-triangle"></i>
                    No se encontraron datos de trazabilidad
                </h4>
                <p>
                    No se encontraron datos de trazabilidad para el producto
                    <strong>%s</strong> (Lote: <strong>%s</strong>)
                    en el período seleccionado.
                </p>
                <hr/>
                <p class="mb-0">
                    <strong>Verifique que:</strong>
                </p>
                <ul>
                    <li>El producto y lote sean correctos</li>
                    <li>Existan partes de producción en el período indicado</li>
                    <li>Las materias primas hayan sido registradas correctamente</li>
                </ul>
            """ % (self.product_id.name, self.lot_id.name)

        return {
            "type": "ir.actions.act_window",
            "res_model": "cdv.production.trace.report",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
            "context": self.env.context,
        }

    def action_print_report(self):
        """Imprimir informe PDF"""
        self.ensure_one()

        if not self.line_ids:
            raise UserError(
                _("Debe calcular la trazabilidad antes de imprimir el informe.")
            )

        return self.env.ref(
            "cdv_production_traceability.action_report_traceability"
        ).report_action(self)

    def action_export_excel(self):
        """Exportar a Excel usando openpyxl"""
        self.ensure_one()

        if not self.line_ids:
            raise UserError(_("Debe calcular la trazabilidad antes de exportar."))

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_("La librería openpyxl no está instalada en el servidor."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        summary_font = Font(bold=True, size=10)
        summary_fill = PatternFill("solid", fgColor="DCE6F1")
        thin = Side(border_style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        wrap = Alignment(wrap_text=True, vertical="center")

        # --- Título ---
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = "Informe de Producto Elaborado"
        title_cell.font = Font(bold=True, size=14, color="1F3864")
        title_cell.alignment = center
        ws.row_dimensions[1].height = 24

        # --- Filtros y Resumen ---
        ws.append([])
        ws.append(["FILTROS", "", "", "RESUMEN", "", "", ""])
        ws["A3"].font = summary_font
        ws["D3"].font = summary_font

        product_name = self.product_id.display_name if self.product_id else "Todos"
        lot_name = self.lot_id.name if self.lot_id else "Todos"

        rows_info = [
            ("Fecha desde:", str(self.date_from), "", "Total producciones:", self.total_productions, "", ""),
            ("Fecha hasta:", str(self.date_to), "", "Total unidades producidas:", f"{self.total_qty_produced} {self.total_qty_uom_name or ''}".strip(), "", ""),
            ("Producto:", product_name, "", "Total componentes rastreados:", self.total_components, "", ""),
            ("Lote:", lot_name, "", "", "", "", ""),
        ]
        for row_data in rows_info:
            ws.append(list(row_data))
            row_num = ws.max_row
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=10)
            ws.cell(row=row_num, column=4).font = Font(bold=True, size=10)

        ws.append([])

        # --- Cabecera de tabla ---
        header_row = [
            "Fecha",
            "Producto Terminado",
            "Lote Producto",
            "Cantidad Producida",
            "Materia Prima",
            "Lote Materia Prima",
            "Estado",
        ]
        ws.append(header_row)
        header_row_num = ws.max_row
        for col_idx, _ in enumerate(header_row, 1):
            cell = ws.cell(row=header_row_num, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # --- Filas de datos ---
        fill_found = PatternFill("solid", fgColor="E2EFDA")
        fill_missing = PatternFill("solid", fgColor="FCE4D6")

        for line in self.line_ids:
            status_text = "Encontrado" if line.trace_status == "found" else "No encontrado"
            row_vals = [
                str(line.production_date) if line.production_date else "",
                line.finished_product_id.display_name if line.finished_product_id else "",
                line.finished_lot_id.name if line.finished_lot_id else "",
                f"{line.finished_qty} {line.finished_uom_id.name if line.finished_uom_id else ''}".strip(),
                line.component_product_id.display_name if line.component_product_id else "",
                line.component_lot_id.name if line.component_lot_id else "",
                status_text,
            ]
            ws.append(row_vals)
            row_num = ws.max_row
            row_fill = fill_found if line.trace_status == "found" else fill_missing
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = row_fill
                cell.border = border
                cell.alignment = wrap
            # Cantidad con formato numérico
            ws.cell(row=row_num, column=4).number_format = "0.00"
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="right", vertical="center")

        # --- Ajustar anchos de columna ---
        col_widths = [14, 28, 18, 18, 28, 18, 16]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # --- Guardar en buffer y crear adjunto ---
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_data = base64.b64encode(buffer.read())

        filename = "trazabilidad_%s_%s.xlsx" % (
            (self.product_id.default_code or self.product_id.name or "producto").replace(" ", "_"),
            self.lot_id.name.replace(" ", "_") if self.lot_id else "todos",
        )

        attachment = self.env["ir.attachment"].create({
            "name": filename,
            "type": "binary",
            "datas": file_data,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%d?download=true" % attachment.id,
            "target": "new",
        }

    def action_reset(self):
        """Volver a borrador para nueva búsqueda"""
        self.ensure_one()
        self.line_ids.unlink()
        self.warning_message = False
        return {
            "type": "ir.actions.act_window",
            "res_model": "cdv.production.trace.report",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
            "context": self.env.context,
        }


class CdvProductionTraceReportLine(models.TransientModel):
    _name = "cdv.production.trace.report.line"
    _description = "Línea de informe de producto elaborado"
    _order = "production_date desc, finished_product_id, component_product_id"

    wizard_id = fields.Many2one(
        comodel_name="cdv.production.trace.report",
        string="Wizard",
        required=True,
        ondelete="cascade",
    )
    production_date = fields.Date(
        string="Fecha producción",
        required=True,
        index=True,
    )
    finished_product_id = fields.Many2one(
        comodel_name="product.product",
        string="Producto terminado",
        required=True,
        index=True,
    )
    finished_lot_id = fields.Many2one(
        comodel_name="stock.lot",
        string="Lote producto terminado",
        index=True,
    )
    finished_qty = fields.Float(
        string="Cantidad producida",
        digits="Product Unit of Measure",
    )
    finished_uom_id = fields.Many2one(
        comodel_name="uom.uom",
        string="UdM producto terminado",
    )
    component_product_id = fields.Many2one(
        comodel_name="product.product",
        string="Componente (materia prima)",
        index=True,
    )
    component_qty_standard = fields.Float(
        string="Cantidad estándar componente",
        digits="Product Unit of Measure",
        help="Cantidad del componente según BoM",
    )
    component_lot_id = fields.Many2one(
        comodel_name="stock.lot",
        string="Lote componente",
        index=True,
    )
    trace_status = fields.Selection(
        selection=[
            ("found", "Encontrado"),
            ("missing", "No encontrado"),
        ],
        string="Estado trazabilidad",
        required=True,
        index=True,
    )
    trace_status_display = fields.Char(
        string="Estado",
        compute="_compute_trace_status_display",
    )

    @api.depends("trace_status")
    def _compute_trace_status_display(self):
        for record in self:
            if record.trace_status == "found":
                record.trace_status_display = "✓ Encontrado"
            else:
                record.trace_status_display = "⚠ No encontrado"
